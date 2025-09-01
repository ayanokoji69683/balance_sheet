import streamlit as st
import pandas as pd
import pdfplumber
import re
from io import BytesIO
import openpyxl
import time
from datetime import datetime
import threading
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor, as_completed
import numpy as np
import google.generativeai as genai

# Set up Gemini API
def setup_gemini(api_key):
    try:
        genai.configure(api_key=api_key)
        return genai
    except Exception as e:
        st.error(f"Error setting up Gemini client: {e}")
        return None

# Function to detect and preserve non-monetary content - FIXED VERSION
def is_non_monetary_content(text):
    """
    Detect if text contains dates, CIN numbers, DIN numbers, phone numbers, 
    FIRM NO., MEM NO. or other non-monetary content
    """
    if not isinstance(text, str) or not text.strip():
        return False
    
    text_lower = text.lower().strip()
    
    # Quick checks first - formulas
    if text_lower.startswith('='):
        return True
    
    # Check if it's purely a year (like 2025, 2024, etc.)
    if re.match(r'^\d{4}$', text):
        return True
    
    # Check if it's a financial year pattern (FY2025, 2024-25, etc.)
    if re.match(r'^(FY\d{4}|\d{4}-\d{2})$', text, re.IGNORECASE):
        return True
    
    # Check for simple numeric patterns that don't need Gemini
    # Allow decimal numbers with proper formatting
    if re.match(r'^-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?$', text):
        # But exclude dates that look like numbers (like 31.03.2025)
        if re.search(r'\d{1,2}\.\d{1,2}\.\d{4}', text):
            return True
        return False
    
    # Date patterns - IMPROVED
    date_patterns = [
        r'\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4}',  # DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY
        r'\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}',    # YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD
        r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]* \d{1,2},? \d{4}',
        r'\d{1,2}(st|nd|rd|th) (jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]* \d{4}',
        r'\d{1,2}\.\d{1,2}\.\d{4}',  # Specific pattern for 31.03.2025
        r'as at \d{1,2}[-/\.]\d{1,2}[-/\.]\d{4}',  # "as at 31.03.2025"
        r'closing.*\d{4}',  # "Closing WDV as at 2025"
    ]
    
    for pattern in date_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    
    # Check for date-like phrases
    date_phrases = [
        'as at', 'at', 'on', 'date', 'year', 'period', 'closing',
        'opening', 'beginning', 'end', 'financial year', 'fy'
    ]
    
    if any(phrase in text_lower for phrase in date_phrases) and any(char.isdigit() for char in text):
        return True
    
    # Other patterns
    patterns = [
        r'[A-Z]{1}[0-9]{5}[A-Z]{2}[0-9]{4}[A-Z]{3}[0-9]{6}',  # CIN
        r'[A-Z]{3}[0-9]{5}',  # DIN-like
        r'[+]{0,1}[0-9]{2,4}[- ]{0,1}[0-9]{3}[- ]{0,1}[0-9]{3}[- ]{0,1}[0-9]{4}',  # Phone
        r'^\d{1,2}(st|nd|rd|th)$',  # Day suffix
        r'(mem|firm|reg|id|no)[\. ]*\d+',  # MEM NO., FIRM NO., etc.
    ]
    
    for pattern in patterns:
        if re.search(pattern, text, re.IGNORECASE):
            return True
    
    # Days and months
    days_months = [
        "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
        "january", "february", "march", "april", "may", "june", "july", "august",
        "september", "october", "november", "december", "jan", "feb", "mar", "apr",
        "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"
    ]
    
    if text_lower in days_months:
        return True
    
    # Check for text that contains both words and numbers but is not monetary
    words = text_lower.split()
    if len(words) > 1 and any(word.isalpha() for word in words) and any(word.isdigit() for word in words):
        # If it contains date-related words with numbers, preserve it
        date_related_words = ['year', 'date', 'period', 'closing', 'opening', 'as', 'at', 'on']
        if any(word in date_related_words for word in words):
            return True
    
    return False

# Cache Gemini responses
@lru_cache(maxsize=1000)
def cached_gemini_extraction(text, api_key):
    """Cached version of Gemini extraction"""
    if not text or not any(char.isdigit() for char in str(text)):
        return []
    
    # Skip if it's clearly non-monetary content
    if is_non_monetary_content(text):
        return []
    
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel('gemini-2.0-flash')
        
        prompt = f"""Extract only monetary values from this text: "{text}". 
        Return only the numbers separated by commas. 
        Ignore dates, years, phone numbers, IDs, and other non-monetary values.
        Example: "Closing balance as at 31.03.2025 is 1,50,000" should return "150000"
        Example: "Year 2025 revenue" should return nothing"""
        
        response = model.generate_content(prompt)
        
        if response and response.text:
            numbers_text = response.text.strip()
            numbers = []
            for part in numbers_text.split(','):
                clean_part = re.sub(r'[^\d.-]', '', part.strip())
                if clean_part and clean_part.replace('.', '', 1).replace('-', '', 1).isdigit():
                    numbers.append(float(clean_part))
            return numbers
    except Exception as e:
        print(f"Gemini API error: {e}")
        pass
    
    return []

# Batch processing function - UPDATED
def process_cell_batch(cell_values, conversion_unit, threshold, api_key=None):
    """Process a batch of cells efficiently"""
    results = []
    conversion_factors = {
        "Hundred": 100,
        "Thousand": 1000,
        "Lakhs": 100000,
        "Crore": 10000000
    }
    factor = conversion_factors[conversion_unit]
    
    for val in cell_values:
        try:
            # Skip processing for non-monetary content - CHECK FIRST
            if isinstance(val, str) and is_non_monetary_content(val):
                results.append(val)
                continue
            
            # Handle numeric values quickly
            if isinstance(val, (int, float)):
                if abs(val) > threshold:
                    converted = val / factor
                    results.append(int(converted) if converted.is_integer() else round(converted, 2))
                else:
                    results.append(val)
                continue
            
            # Handle string values
            if isinstance(val, str):
                # Quick regex extraction for simple cases
                num_matches = re.findall(r'-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?', val.replace(',', ''))
                if num_matches:
                    numbers = [float(match.replace(',', '')) for match in num_matches]
                    
                    # Check if this might be a date disguised as numbers
                    if len(numbers) == 3 and all(0 < num < 32 for num in numbers[:2]) and numbers[2] > 1900:
                        results.append(val)
                        continue
                    
                    largest_num = max(numbers, key=abs) if numbers else 0
                    
                    if abs(largest_num) > threshold:
                        converted_num = largest_num / factor
                        if converted_num.is_integer():
                            converted_num = int(converted_num)
                        else:
                            converted_num = round(converted_num, 2)
                        
                        # Replace the largest number found
                        new_val = val
                        for num in sorted(numbers, key=lambda x: len(str(x)), reverse=True):
                            if abs(num) > threshold:
                                converted = num / factor
                                if converted.is_integer():
                                    converted = int(converted)
                                else:
                                    converted = round(converted, 2)
                                new_val = new_val.replace(str(num), str(converted))
                        results.append(new_val)
                    else:
                        results.append(val)
                else:
                    # Use Gemini only for complex cases with numbers
                    if any(char.isdigit() for char in val) and api_key:
                        numbers = cached_gemini_extraction(val, api_key)
                        if numbers:
                            largest_num = max(numbers, key=abs)
                            if abs(largest_num) > threshold:
                                converted_num = largest_num / factor
                                converted_num = int(converted_num) if converted_num.is_integer() else round(converted_num, 2)
                                results.append(str(converted_num))
                            else:
                                results.append(val)
                        else:
                            results.append(val)
                    else:
                        results.append(val)
            else:
                results.append(val)
                
        except Exception:
            results.append(val)
    
    return results

def convert_units_in_cell(val, conversion_unit="Lakhs", threshold=20, api_key=None):
    """Optimized single cell conversion"""
    conversion_factors = {
        "Hundred": 100,
        "Thousand": 1000,
        "Lakhs": 100000,
        "Crore": 10000000
    }
    factor = conversion_factors[conversion_unit]
    
    try:
        # Skip processing for non-monetary content
        if isinstance(val, str) and is_non_monetary_content(val):
            return val
        
        # Handle numeric values
        if isinstance(val, (int, float)):
            if abs(val) > threshold:
                converted = val / factor
                return int(converted) if converted.is_integer() else round(converted, 2)
            return val
        
        # Handle string values
        if isinstance(val, str):
            # Quick regex extraction
            num_matches = re.findall(r'-?\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?', val.replace(',', ''))
            if num_matches:
                numbers = [float(match.replace(',', '')) for match in num_matches]
                
                # Check if this might be a date
                if len(numbers) == 3 and all(0 < num < 32 for num in numbers[:2]) and numbers[2] > 1900:
                    return val
                
                largest_num = max(numbers, key=abs) if numbers else 0
                
                if abs(largest_num) > threshold:
                    # Replace all convertible numbers
                    new_val = val
                    for num in sorted(numbers, key=lambda x: len(str(x)), reverse=True):
                        if abs(num) > threshold:
                            converted = num / factor
                            if converted.is_integer():
                                converted = int(converted)
                            else:
                                converted = round(converted, 2)
                            new_val = new_val.replace(str(num), str(converted))
                    return new_val
            return val
        
        return val
    except Exception:
        return val

def add_unit_row(df, conversion_unit):
    """Add an extra first row showing units only for columns with converted numeric values."""
    unit_row = []
    for col in df.columns:
        # Sample first few rows to check for numeric values
        sample_values = df[col].head(10).dropna()
        has_converted = any(
            isinstance(val, (int, float)) and abs(val) < 1000 and val == round(val, 2)
            for val in sample_values
        )
        unit_row.append(f"(in {conversion_unit})" if has_converted else "")
    
    df_with_units = pd.DataFrame([unit_row], columns=df.columns)
    df_with_units = pd.concat([df_with_units, df], ignore_index=True)
    return df_with_units

def extract_tables_from_pdf(file_bytes, conversion_unit, api_key=None):
    """Optimized PDF table extraction"""
    all_tables = {}
    
    with pdfplumber.open(file_bytes) as pdf:
        total_pages = len(pdf.pages)
        
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if tables:
                for j, table in enumerate(tables):
                    df = pd.DataFrame(table)
                    
                    # Process entire DataFrame at once
                    for col in df.columns:
                        df[col] = process_cell_batch(df[col].tolist(), conversion_unit, 20, api_key)
                    
                    df = add_unit_row(df, conversion_unit)
                    all_tables[f"Page_{i+1}_Table_{j+1}"] = df
            
            # Update progress
            progress = (i + 1) / total_pages
            st.session_state.progress_bar.progress(progress)
            st.session_state.status_text.text(f"Processing PDF page {i+1}/{total_pages}")
    
    return all_tables

def create_preserve_excel(excel_bytes, conversion_unit, api_key=None):
    """Optimized Excel processing"""
    # Load workbook once for data
    wb_data = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=True)
    wb = openpyxl.load_workbook(BytesIO(excel_bytes), data_only=False)
    
    total_sheets = len(wb.sheetnames)
    total_cells = sum(ws.max_row * ws.max_column for ws in wb_data.worksheets)
    processed_cells = 0
    
    for sheet_idx, ws_name in enumerate(wb.sheetnames):
        ws_data = wb_data[ws_name]
        ws = wb[ws_name]
        
        # Collect all cell values for batch processing
        cell_values = []
        cell_positions = []
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_data = ws_data.cell(row=row, column=col)
                cell = ws.cell(row=row, column=col)
                
                if cell.value is not None:
                    cell_values.append(cell_data.value if cell_data.value is not None else cell.value)
                    cell_positions.append((row, col))
                processed_cells += 1
                
                # Update progress periodically
                if processed_cells % 100 == 0:
                    progress = processed_cells / total_cells
                    st.session_state.progress_bar.progress(progress)
                    st.session_state.status_text.text(f"Processing cell {processed_cells}/{total_cells}")
        
        # Process cells in batches
        processed_values = process_cell_batch(cell_values, conversion_unit, 20, api_key)
        
        # Update cells with processed values
        for (row, col), new_value in zip(cell_positions, processed_values):
            ws.cell(row=row, column=col).value = new_value
        
        # Add unit row
        unit_row = []
        for col in range(1, ws.max_column + 1):
            has_numeric = any(
                isinstance(ws.cell(row=row, column=col).value, (int, float))
                for row in range(2, min(ws.max_row + 1, 100))  # Check first 100 rows
            )
            unit_row.append(f"(in {conversion_unit})" if has_numeric else "")
        
        ws.insert_rows(1)
        for idx, val in enumerate(unit_row, start=1):
            ws.cell(row=1, column=idx).value = val
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.set_page_config(page_title="Balance Sheet Converter", layout="wide")
st.title("📊 Balance Sheet Converter with Gemini 2.0 Flash")

# Initialize session state
if 'progress_bar' not in st.session_state:
    st.session_state.progress_bar = st.empty()
if 'status_text' not in st.session_state:
    st.session_state.status_text = st.empty()

# Gemini API key input
api_key = st.sidebar.text_input("Enter Gemini API Key:", type="password")
use_gemini = bool(api_key)

if api_key:
    try:
        genai_client = setup_gemini(api_key)
        if genai_client:
            st.sidebar.success("Gemini API configured successfully!")
        else:
            st.sidebar.error("Failed to configure Gemini API")
            use_gemini = False
    except Exception as e:
        st.sidebar.error(f"Error configuring Gemini API: {e}")
        use_gemini = False
else:
    st.sidebar.warning("Add a Gemini API key for enhanced conversion (optional)")

st.sidebar.info("""
This tool converts monetary values in balance sheets while preserving:
- Dates and time periods (2025, 31.03.2025, etc.)
- CIN numbers and identifiers  
- Phone numbers
- Formulas and calculations
- Text descriptions
- Years and financial periods
""")

conversion_unit = st.selectbox(
    "Select conversion unit:",
    ["Crore", "Lakhs", "Thousand", "Hundred"],
    index=1
)

threshold = st.slider(
    "Conversion threshold (values below this won't be converted):",
    min_value=0,
    max_value=100,
    value=20
)

uploaded_file = st.file_uploader("Choose an Excel or PDF file", type=["xlsx", "xls", "pdf"])

if uploaded_file is not None:
    file_type = uploaded_file.name.split('.')[-1].lower()
    file_bytes = uploaded_file.read()

    if file_type in ["xlsx", "xls"]:
        st.success(f"Processing Excel file: {uploaded_file.name}")

        # Show original preview
        try:
            original_df = pd.read_excel(BytesIO(file_bytes))
            st.subheader("Original Values Preview")
            st.dataframe(original_df.head(6))
        except Exception as e:
            st.warning(f"Could not display original file preview: {e}")

        # Process file
        st.session_state.progress_bar = st.progress(0)
        st.session_state.status_text = st.empty()
        
        excel_output = create_preserve_excel(file_bytes, conversion_unit, api_key if use_gemini else None)
        
        st.session_state.progress_bar.progress(1.0)
        st.session_state.status_text.text("Conversion complete!")
        time.sleep(0.5)
        st.session_state.progress_bar.empty()
        st.session_state.status_text.empty()

        # Show converted preview
        try:
            converted_df = pd.read_excel(excel_output)
            st.subheader("Converted Values Preview")
            st.dataframe(converted_df.head(6))
            
            # Statistics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Columns", len(converted_df.columns))
            with col2:
                st.metric("Rows", len(converted_df))
            with col3:
                numeric_cols = sum([converted_df[col].apply(lambda x: isinstance(x, (int, float))).any() 
                                  for col in converted_df.columns])
                st.metric("Numeric Columns", numeric_cols)
                
        except Exception as e:
            st.warning(f"Could not display converted file preview: {e}")

        # Download button
        st.download_button(
            label=f"📥 Download Converted Excel",
            data=excel_output,
            file_name=f"converted_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    elif file_type == "pdf":
        st.success(f"Processing PDF file: {uploaded_file.name}")
        
        # Process PDF
        st.session_state.progress_bar = st.progress(0)
        st.session_state.status_text = st.empty()
        
        tables_dict = extract_tables_from_pdf(BytesIO(file_bytes), conversion_unit, api_key if use_gemini else None)
        
        st.session_state.progress_bar.progress(1.0)
        st.session_state.status_text.text("Processing completed!")
        time.sleep(0.5)
        st.session_state.progress_bar.empty()
        st.session_state.status_text.empty()

        if tables_dict:
            st.write(f"**Found {len(tables_dict)} tables in the PDF**")
            
            # Navigation
            table_names = list(tables_dict.keys())
            selected_table = st.selectbox("Select table to view:", table_names)
            
            st.write(f"**{selected_table}**")
            st.dataframe(tables_dict[selected_table].head(8))

            # Create Excel output
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for sheet_name, df in tables_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False, header=False)
            output.seek(0)

            st.download_button(
                label=f"📥 Download Excel",
                data=output,
                file_name=f"converted_{uploaded_file.name.split('.')[0]}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("❌ No tables found in this PDF.")

    # Conversion examples
    with st.expander("ℹ️ See conversion examples"):
        st.write("""
        **Examples of values that will be converted:**
        - `1,50,000` → `1.5` (in Lakhs)
        - `2500000` → `25` (in Lakhs)
        - `75000000` → `7.5` (in Crore)
        
        **Examples of values that will be PRESERVED (not converted):**
        - Years: `2025`, `2024`, `FY2025`
        - Dates: `31.03.2025`, `31-03-2025`, `March 31, 2025`
        - Date phrases: `Closing WDV as at 31.03.2025`, `As at 2025`
        - CIN Numbers: `U72300DL2015NPL285463`
        - DIN Numbers: `01234567`
        - MEM NO. : `432522`    
        - FIRM NO. : `021992C`   
        - Phone Numbers: `+91-1234567890`
        - Formulas: `=SUM(A1:A10)`
        - Text: `Authorised Capital`
        """)
    # API status
    if use_gemini:
        st.sidebar.success("Using Gemini 2.0 Flash for enhanced conversion")
    else:
        st.sidebar.info("Using regex-based conversion (add API key for better results)")